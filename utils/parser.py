"""
parser.py — SENKA/UNO 리포트 엑셀 파싱 모듈
엑셀 구조를 자동 감지하여 소재별 성과 데이터를 추출합니다.
"""
import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Tuple
import re


def parse_report(file) -> Dict:
    """메인 파싱 함수: 엑셀 파일을 받아 구조화된 데이터를 반환"""
    wb = openpyxl.load_workbook(file, data_only=True)
    
    result = {
        "meta": parse_summary(wb),
        "creatives": parse_creatives(wb),
        "daily": parse_daily(wb),
        "media_detail": parse_media_detail(wb),
    }
    
    return result


def parse_summary(wb) -> Dict:
    """Summary 시트에서 캠페인 메타 정보 추출"""
    if "Summary" not in wb.sheetnames:
        return {}
    
    ws = wb["Summary"]
    meta = {}
    
    # 기본 정보 (B2~B6 영역)
    field_map = {
        "Client": "client",
        "Campaign": "campaign", 
        "Reporting Period": "period",
        "Campaign Budget": "budget",
        "Site": "site",
    }
    
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=2, max_col=7, values_only=False):
        for cell in row:
            if cell.value and str(cell.value).strip() in field_map:
                key = field_map[str(cell.value).strip()]
                # C열에서 값 가져오기
                val_cell = ws.cell(row=cell.row, column=cell.column + 1)
                meta[key] = val_cell.value
    
    # Total 성과 지표 (B16 행)
    meta["total_performance"] = {}
    for row in ws.iter_rows(min_row=10, max_row=40, values_only=False):
        row_vals = {c.column: c.value for c in row if c.value is not None}
        if 2 in row_vals and row_vals[2] == "TOTAL":
            meta["total_performance"] = {
                "impressions": row_vals.get(5, 0),
                "clicks": row_vals.get(8, 0),
                "actions": row_vals.get(11, 0),
                "ctr": row_vals.get(14, 0),
                "spent_budget": row_vals.get(16, 0),
                "cpm": row_vals.get(19, 0),
                "cpc": row_vals.get(21, 0),
            }
            break
    
    # 제품별 소계 파싱
    products = {}
    current_product = None
    for row in ws.iter_rows(min_row=17, max_row=50, values_only=False):
        row_dict = {c.column: c.value for c in row if c.value is not None}
        # 제품명 행 감지 (B열에 제품명만 있는 행)
        if 2 in row_dict and len(row_dict) <= 2 and isinstance(row_dict[2], str):
            current_product = row_dict[2]
            products[current_product] = {}
        # TOTAL 행 감지
        if 2 in row_dict and row_dict[2] == "TOTAL" and current_product:
            products[current_product] = {
                "impressions": row_dict.get(5, 0),
                "clicks": row_dict.get(8, 0),
                "actions": row_dict.get(11, 0),
                "spent_budget": row_dict.get(16, 0),
            }
    
    meta["products"] = products
    return meta


def parse_creatives(wb) -> List[Dict]:
    """Creative 시트에서 소재별 성과 데이터 추출"""
    if "Creative" not in wb.sheetnames:
        return []
    
    ws = wb["Creative"]
    creatives = []
    
    # META(전환_구매) 소재 파싱 (B24~ 영역)
    current_media = None
    current_product = None
    
    for row_idx in range(24, ws.max_row + 1):
        row = {}
        for col_idx in range(2, 21):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                row[col_idx] = cell.value
        
        if not row:
            continue
        
        # MEDIA 헤더 감지 (B열)
        if 2 in row and isinstance(row[2], str):
            val = str(row[2]).strip()
            if val.startswith("META") or val.startswith("구글") or val == "Grand TOTAL" or val == "SUB TOTAL":
                current_media = val
                if current_media in ("Grand TOTAL", "SUB TOTAL"):
                    continue
        
        # 제품명 감지 (C열)
        if 3 in row and isinstance(row[3], str) and row[3].strip():
            current_product = row[3].strip()
        
        # 소재 데이터 행 (D열에 "N안" 패턴)
        if 4 in row and isinstance(row[4], str):
            name = str(row[4]).strip()
            if not name:
                continue
            
            creative = {
                "media": current_media or "",
                "product": current_product or "",
                "name": name,
                "impressions": _safe_num(row.get(6)),
                "clicks": _safe_num(row.get(7)),
                "actions": _safe_num(row.get(8)),
                "revenue": _safe_num(row.get(9)),
                "cvr": _safe_num(row.get(10)),
                "ctr": _safe_num(row.get(11)),
                "cpm": _safe_num(row.get(12)),
                "cpc": _safe_num(row.get(13)),
                "cpa": _safe_num(row.get(14)),
                "roas": _safe_num(row.get(15)),
                "spent": _safe_num(row.get(16)),
                "note": str(row.get(17, "")) if row.get(17) else "",
            }
            
            # 소재 유형 판별 (이미지/영상/카탈로그)
            if "이미지" in name:
                creative["format"] = "이미지"
            elif "영상" in name:
                creative["format"] = "영상"
            elif "카탈로그" in name:
                creative["format"] = "카탈로그"
            elif "디멘드젠" in name or "디맨드젠" in name:
                creative["format"] = "디멘드젠"
            else:
                creative["format"] = "기타"
            
            creatives.append(creative)
    
    return creatives


def parse_daily(wb) -> pd.DataFrame:
    """Daily 시트에서 일별 성과 데이터 추출"""
    if "Daily" not in wb.sheetnames:
        return pd.DataFrame()
    
    ws = wb["Daily"]
    rows = []
    
    for row_idx in range(13, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=2).value
        if date_cell is None:
            continue
        
        row = {
            "date": date_cell,
            "day_of_week": ws.cell(row=row_idx, column=3).value,
            # Total
            "total_imps": _safe_num(ws.cell(row=row_idx, column=4).value),
            "total_clicks": _safe_num(ws.cell(row=row_idx, column=5).value),
            "total_actions": _safe_num(ws.cell(row=row_idx, column=6).value),
            "total_ctr": _safe_num(ws.cell(row=row_idx, column=7).value),
            "total_cost": _safe_num(ws.cell(row=row_idx, column=11).value),
            # META 전환
            "meta_conv_imps": _safe_num(ws.cell(row=row_idx, column=15).value),
            "meta_conv_clicks": _safe_num(ws.cell(row=row_idx, column=16).value),
            "meta_conv_actions": _safe_num(ws.cell(row=row_idx, column=17).value),
            "meta_conv_cost": _safe_num(ws.cell(row=row_idx, column=22).value),
            # META 트래픽
            "meta_traffic_imps": _safe_num(ws.cell(row=row_idx, column=23).value),
            "meta_traffic_clicks": _safe_num(ws.cell(row=row_idx, column=24).value),
            "meta_traffic_cost": _safe_num(ws.cell(row=row_idx, column=30).value),
            # 구글
            "google_imps": _safe_num(ws.cell(row=row_idx, column=31).value),
            "google_clicks": _safe_num(ws.cell(row=row_idx, column=32).value),
            "google_cost": _safe_num(ws.cell(row=row_idx, column=36).value),
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    if not df.empty and "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"])
    return df


def parse_media_detail(wb) -> Dict:
    """META(전환), META(트래픽), 구글 시트에서 제품별 일별 상세 추출"""
    detail = {}
    
    # META(전환) 시트 — 제품별 daily
    if "META(전환)" in wb.sheetnames:
        ws = wb["META(전환)"]
        detail["meta_conversion"] = _parse_media_sheet_daily(ws)
    
    if "META(트래픽)" in wb.sheetnames:
        ws = wb["META(트래픽)"]
        detail["meta_traffic"] = _parse_media_sheet_daily(ws)
    
    return detail


def _parse_media_sheet_daily(ws) -> Dict:
    """매체 시트에서 Total 성과 + 일별 데이터 추출"""
    result = {
        "total": {},
        "daily": [],
    }
    
    # Total (row 9)
    result["total"] = {
        "impressions": _safe_num(ws.cell(row=9, column=4).value),
        "clicks": _safe_num(ws.cell(row=9, column=5).value),
        "actions": _safe_num(ws.cell(row=9, column=6).value),
        "revenue": _safe_num(ws.cell(row=9, column=7).value),
        "ctr": _safe_num(ws.cell(row=9, column=8).value),
        "cvr": _safe_num(ws.cell(row=9, column=9).value),
        "roas": _safe_num(ws.cell(row=9, column=10).value),
    }
    
    # Daily (row 17~)
    for row_idx in range(17, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=2).value
        if date_val is None:
            continue
        row = {
            "date": date_val,
            "impressions": _safe_num(ws.cell(row=row_idx, column=4).value),
            "clicks": _safe_num(ws.cell(row=row_idx, column=5).value),
            "actions": _safe_num(ws.cell(row=row_idx, column=6).value),
            "revenue": _safe_num(ws.cell(row=row_idx, column=7).value),
            "ctr": _safe_num(ws.cell(row=row_idx, column=8).value),
            "cvr": _safe_num(ws.cell(row=row_idx, column=9).value),
            "roas": _safe_num(ws.cell(row=row_idx, column=10).value),
            "cost": _safe_num(ws.cell(row=row_idx, column=14).value),
        }
        result["daily"].append(row)
    
    return result


def _safe_num(val) -> float:
    """안전한 숫자 변환"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("%", "")
        if val in ("#DIV/0!", "#REF!", "#N/A", "", "-"):
            return 0.0
        try:
            return float(val)
        except ValueError:
            return 0.0
    return 0.0
